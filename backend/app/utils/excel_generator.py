from typing import List, Optional
from datetime import datetime
from decimal import Decimal
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

class ExcelGenerator:
    def __init__(self):
        self.workbook = Workbook()
        self.sheet = self.workbook.active
        self._setup_page()

    def _setup_page(self):
        """A4 사이즈에 맞게 페이지 설정"""
        self.sheet.page_setup.paperSize = self.sheet.PAPERSIZE_A4
        self.sheet.page_setup.orientation = self.sheet.ORIENTATION_PORTRAIT
        
        # A4 크기에 맞게 칼럼 너비 조정
        column_widths = [5, 40, 10, 15, 15, 30]  # 각 칼럼의 너비
        for i, width in enumerate(column_widths, 1):
            self.sheet.column_dimensions[get_column_letter(i)].width = width

    def _apply_header_style(self, cell):
        """헤더 셀 스타일 적용"""
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="CCCCCC")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

    def _apply_cell_style(self, cell, align="left"):
        """일반 셀 스타일 적용"""
        cell.alignment = Alignment(horizontal=align, vertical="center")
        cell.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

    def generate_quotation(self, quotation_data: dict) -> Workbook:
        """견적서 Excel 생성"""
        # 회사 정보 및 견적서 정보
        self.sheet["B1"] = "Quotation"
        self.sheet["B1"].font = Font(size=16, bold=True)
        
        self.sheet["B3"] = f"QUOTE #: {quotation_data['quote_number']}"
        self.sheet["B4"] = f"DATE: {datetime.now().strftime('%Y-%m-%d')}"
        self.sheet["B5"] = f"VALID UNTIL: {quotation_data['valid_until']}"
        
        self.sheet["D3"] = "Customer:"
        self.sheet["E3"] = quotation_data['customer_name']
        self.sheet["D4"] = "Project Description:"
        self.sheet["E4"] = quotation_data['project_description']

        # 견적서 항목 헤더
        headers = ["No.", "Item", "Q'ty", "Unit Price", "Amount", "Remark"]
        for col, header in enumerate(headers, 1):
            cell = self.sheet.cell(row=7, column=col)
            cell.value = header
            self._apply_header_style(cell)

        # 견적서 항목
        current_row = 8
        for idx, item in enumerate(quotation_data['items'], 1):
            cells = [
                (str(idx), "center"),
                (item['name'], "left"),
                (str(item['quantity']), "center"),
                (f"{item['unit_price']:,}", "right"),
                (f"{item['quantity'] * item['unit_price']:,}", "right"),
                (item.get('remark', ''), "left")
            ]
            
            for col, (value, align) in enumerate(cells, 1):
                cell = self.sheet.cell(row=current_row, column=col)
                cell.value = value
                self._apply_cell_style(cell, align)
            
            current_row += 1

        # 합계
        self.sheet.cell(row=current_row + 1, column=1, value="Total")._apply_header_style()
        total_cell = self.sheet.cell(row=current_row + 1, column=5)
        total_cell.value = f"{quotation_data['total_amount']:,}"
        self._apply_cell_style(total_cell, "right")

        # 참고사항
        self.sheet.cell(row=current_row + 3, column=1, value="Note:")
        self.sheet.cell(row=current_row + 4, column=1, value="1. Currency: KRW")
        self.sheet.cell(row=current_row + 5, column=1, value="2. VAT excluded")

        return self.workbook

def format_currency(amount: Decimal) -> str:
    """통화 형식으로 변환"""
    return f"{amount:,.0f}"